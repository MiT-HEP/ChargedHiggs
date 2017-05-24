import sys,os
import openpyxl

class Read:
    def __init__(self):

        self.columns_map  = {  
                ('YR3','ggH') : 'B',
                ('YR3','qqH') : 'I',
                ('YR3','WPlusH') : 'U',
                ('YR3','WMinusH') : 'V',
                ('YR3','WH'):'P',
                ('YR3','ZH') : 'Y',
                ('YR3','ttH') : 'AG',
                ('YR3','bbH') : 'AN',

                ('YR3','Hmm') : 'H',  ## BR

                ('YR4','ggH') : 'B',
                ('YR4','qqH') : 'S',
                ('YR4','WPlusH') : 'AG',
                ('YR4','WMinusH') : 'AH',
                ('YR4','WH'):'AA',
                ('YR4','ZH') : 'AK',
                ('YR4','ttH') : 'AT',
                ('YR4','bbH') : 'BB',

                ('YR4','Hmm') : 'P',  ## BR
                }

        self.sheet_map  = {  
                ('YR3','BR' ) : 'BR',
                ('YR4','BR' ) : 'YR4 SM BR',

                ('YR3','7TeV' ) : 'SM 7TeV',
                ('YR4','7TeV' ) : 'YR4 SM 7TeV',

                ('YR3','8TeV' ) : 'SM 8TeV',
                ('YR4','8TeV' ) : 'YR4 SM 8TeV',

                ('YR4','13TeV' ) : 'YR4 SM 13TeV',
                ('YR4','14TeV' ) : 'YR4 SM 14TeV',
                }
        return

    def init(self,fNameYR3="",fNameYR4=""):
        self.wb={}
        self.wb['YR3']  = openpyxl.load_workbook(fNameYR3, read_only=True,data_only=True) ## data_only= read values and not formulas
        self.wb['YR4']  = openpyxl.load_workbook(fNameYR4, read_only=True,data_only=True)
        self.fNames={
                'YR3':fNameYR3,
                'YR4':fNameYR4,
                }
        return

    def readSinglePoint(self,YR='YR4',Ecm='7TeV' ,proc='ggH',mass=125.09 ):
        ''' Ecm = 7TeV/8TeV/BR.
            proc = ggH,qqH,WPlusH,WMinusH,WH,ZH,ttH,bbH -- for 7TeV/8TeV
            proc = Hmm for BR
        '''
        #wb = openpyxl.load_workbook(file_name, read_only=True)
        sheet=self.sheet_map[(YR,Ecm)]
        col = self.columns_map[(YR,proc)]
        try:
            idx= self.wb[YR].get_sheet_names().index(sheet)
        except ValueError:
            print "Unable to find sheet",sheet,"in file",self.fNames[YR]
            print "Available names are:",self.wb[YR].get_sheet_names()
            raise ValueError
        ws = self.wb[YR].worksheets[idx]
        row=-1
        masses=[]
        for irow in range(6,300):
            try:
                masses.append(float(ws['A%d'%irow].value) )
                if abs(float(ws['A%d'%irow].value) - mass) <0.001:
                    row = irow
                    break
            except TypeError: ## YR4 BR starts from row7
                print "Unable to convert to float cell:",'A%d'%irow,": '"+ws['A%d'%irow].value+"'"
                pass
            except ValueError: ## YR4 BR starts from row7
                print "Unable to convert to float cell:",'A%d'%irow,": '"+ws['A%d'%irow].value+"'"
                pass
        if row < 0 : 
            print "Unable to Find mass",mass,"in sheet",sheet,"and file",self.fNames[YR],"("+YR+")"
            print "available masses are:", ",".join(["%.2f"%x for x in masses])
            raise ValueError

        try:
            return float(ws['%s%d'%(col,row)].value)
        except ValueError as e:
            print "Exception: maybe the header helps ... and rows around",row
            print "------------------------------"
            print "|",ws['%s%d'%(col,2)].value,"|"
            print "|",ws['%s%d'%(col,3)].value,"|"
            print "|",ws['%s%d'%(col,4)].value,"|"
            print "|",ws['%s%d'%(col,5)].value,"|"
            print "------------------------------"
            print "|",ws['%s%d'%(col,row-2)].value,"|"
            print "|",ws['%s%d'%(col,row-1)].value,"|"
            print "|",ws['%s%d'%(col,row)].value,"|"
            print "|",ws['%s%d'%(col,row+1)].value,"|"
            print "|",ws['%s%d'%(col,row+2)].value,"|"
            print "------------------------------"
            raise e


if __name__ == '__main__':
    reader= Read()
    d = '/'.join(sys.argv[0].split('/')[:-1]) + '/'
    reader.init(d+'Higgs_XSBR_YR3_update.xlsx',d+'Higgs_XSBR_YR4_update.xlsx')
    if len(sys.argv) >3 :
        yr4 = reader.readSinglePoint('YR4',sys.argv[1],sys.argv[2],float(sys.argv[3]))
        yr3 = reader.readSinglePoint('YR3',sys.argv[1],sys.argv[2],float(sys.argv[3]))
        print "YR4/YR3 (Ecm=%s,proc=%s,mass=%s)"%(sys.argv[1],sys.argv[2],sys.argv[3]),yr4,yr3,'==',yr4/yr3
    else:
        print "Usage:"
        print sys.argv[0],"Ecm proc mass"
        print " Ecm = 7TeV/8TeV/BR "
        print " proc = ggH,qqH,WPlusH,WMinusH,ZH,ttH,bbH -- Hmm for BR"
